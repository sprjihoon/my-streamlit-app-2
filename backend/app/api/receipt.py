"""
backend/app/api/receipt.py - 동대문 장끼/영수증 AI 분석 API
──────────────────────────────────────────────────────────
GPT-4o Vision으로 장끼/영수증 이미지를 분석하여 구조화된 데이터 추출.
ddm/compfirm-mvp 프로젝트의 aiParse.service.ts 로직 Python 포팅.

기능:
- 이미지 업로드 → GPT-4o Vision OCR
- 거래처명, 날짜, 품목(품명/옵션/단가/수량/금액) 자동 추출
- 금액 검산 (단가×수량=금액, 품목합계=영수증합계)
- 확인 필요 항목 자동 표시
- 엑셀 다운로드 (거래일/거래처명/품명/옵션/단가/수량/금액 등)
"""

import base64
import json
import os
import uuid
from datetime import datetime, date
from pathlib import Path
from typing import Optional, List

from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from logic.db import get_connection
from backend.app.config import settings

router = APIRouter(prefix="/receipt", tags=["receipt"])

UPLOAD_DIR = Path(settings.UPLOAD_DIR) / "receipts"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


# ─────────────────────────────────────
# DB 초기화
# ─────────────────────────────────────

def ensure_receipt_tables():
    with get_connection() as con:
        con.execute("""
            CREATE TABLE IF NOT EXISTS receipts (
                id TEXT PRIMARY KEY,
                user_id INTEGER NOT NULL,
                image_filename TEXT,
                store_name TEXT,
                receipt_type TEXT DEFAULT '장끼',
                receipt_no TEXT,
                order_date TEXT,
                phone TEXT,
                total_amount REAL,
                paid_amount REAL,
                balance_amount REAL,
                bank_info TEXT,
                memo TEXT,
                raw_text TEXT,
                image_quality TEXT DEFAULT 'good',
                is_handwritten INTEGER DEFAULT 0,
                confidence REAL DEFAULT 1.0,
                needs_review INTEGER DEFAULT 0,
                warnings TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """)
        con.execute("""
            CREATE TABLE IF NOT EXISTS receipt_items (
                id TEXT PRIMARY KEY,
                receipt_id TEXT NOT NULL,
                line_no INTEGER,
                raw_text TEXT,
                item_name TEXT,
                color TEXT,
                size TEXT,
                option_text TEXT,
                unit_price REAL,
                quantity REAL,
                amount REAL,
                confidence REAL DEFAULT 1.0,
                needs_review INTEGER DEFAULT 0,
                warnings TEXT,
                FOREIGN KEY (receipt_id) REFERENCES receipts(id)
            )
        """)
        con.commit()


# ─────────────────────────────────────
# 인증 헬퍼
# ─────────────────────────────────────

def get_user_from_token(token: str) -> dict:
    with get_connection() as con:
        row = con.execute(
            "SELECT user_id, nickname, is_admin FROM users WHERE session_token = ?",
            (token,)
        ).fetchone()
    if not row:
        raise HTTPException(status_code=401, detail="인증이 필요합니다.")
    return {"user_id": row[0], "nickname": row[1], "is_admin": bool(row[2])}


# ─────────────────────────────────────
# GPT-4o Vision OCR
# ─────────────────────────────────────

SYSTEM_PROMPT = """너는 동대문 장끼/영수증을 엑셀로 정리하는 검수 보조 AI다.
이미지 안의 내용을 단순히 OCR하지 말고,
거래처 정보와 품목 행을 분리해서 구조화된 JSON으로 반환해야 한다.

반드시 아래 JSON 형식만 반환하고, 다른 설명 문장은 반환하지 않는다.

{
  "receipt": {
    "storeName": "거래처명 또는 null",
    "receiptType": "영수증 | 장끼 | 거래명세서",
    "receiptNo": "영수증번호 또는 null",
    "orderDate": "YYYY-MM-DD 형식 또는 null",
    "phone": "연락처 또는 null",
    "totalAmount": 숫자 또는 null,
    "paidAmount": 숫자 또는 null,
    "balanceAmount": 숫자 또는 null,
    "bankInfo": "계좌정보 또는 null",
    "memo": "안내문/메모 또는 null",
    "imageQuality": "good | blurry | dark | tilted | handwritten | mixed",
    "isHandwritten": true | false,
    "confidence": 0.0~1.0,
    "needsReview": true | false,
    "warnings": ["경고 메시지"]
  },
  "items": [
    {
      "lineNo": 1,
      "rawText": "원본 텍스트",
      "itemName": "품명 (색상/사이즈/옵션 제외)",
      "color": "색상 또는 null",
      "size": "사이즈 또는 null",
      "optionText": "기타 옵션 또는 null",
      "unitPrice": 단가 숫자 또는 null,
      "quantity": 수량 숫자 또는 null,
      "amount": 금액 숫자 또는 null,
      "confidence": 0.0~1.0,
      "needsReview": true | false,
      "warnings": ["경고 메시지"]
    }
  ],
  "rawText": "전체 OCR 텍스트",
  "overallWarnings": ["전체 경고"]
}

규칙:
1. 품명과 옵션을 가능하면 분리한다.
   - 품명: 프랑 체크 반셔츠
   - 옵션: 블랙, L
2. 거래처명은 "XX 귀하"가 아니라 발송인/공급자를 찾는다.
3. 수기로 작성된 글씨도 최대한 읽되, 확실하지 않으면 needsReview=true로 표시한다.
4. 잘 안 보이는 글자는 추정값을 넣되 warnings에 "글자 일부 불명확"이라고 남긴다.
5. 전혀 모르겠으면 null로 둔다.
6. 금액은 숫자만 반환한다. 예: "26,000원" → 26000
7. 날짜는 가능하면 YYYY-MM-DD로 정규화한다.
8. 단가 × 수량 = 금액인지 확인한다. 맞지 않으면 warnings에 "금액 불일치"를 남긴다.
9. 품목 금액 합계와 영수증 합계가 맞는지 확인한다.
10. 수기 영수증은 isHandwritten=true, needsReview=true로 표시한다.
11. confidence < 0.8이면 needsReview=true
12. itemName이나 quantity나 amount가 없으면 needsReview=true
13. QR코드, 바코드는 필수로 해석하지 않아도 된다.
14. 하단 안내문, 계좌정보는 memo 또는 bankInfo에 저장한다.

수기 영수증 특별 규칙:
- 수기 글씨에서 숫자는 특히 중요하므로 단가/수량/금액은 별도 검산한다.
- 수기 영수증은 자동 확정하지 않고 반드시 검수 필요 상태로 둔다.
- 불확실한 글자는 물음표와 함께 표시한다. 예: itemName="프랑체크?", warnings=["품명 불확실"]"""


def _get_mime(filename: str) -> str:
    ext = Path(filename).suffix.lower()
    return {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
            "webp": "image/webp", "gif": "image/gif"}.get(ext.lstrip("."), "image/jpeg")


def _validate_amounts(receipt: dict, items: list) -> None:
    """단가×수량=금액 검산, 품목합계=영수증합계 검산"""
    for item in items:
        u = item.get("unitPrice")
        q = item.get("quantity")
        a = item.get("amount")
        if u is not None and q is not None and a is not None:
            calc = u * q
            if abs(calc - a) > 1:
                item["needsReview"] = True
                item.setdefault("warnings", [])
                item["warnings"].append(f"금액 불일치: {u}×{q}={calc} ≠ {a}")
        if item.get("confidence", 1.0) < 0.8:
            item["needsReview"] = True
        if not item.get("itemName") or item.get("quantity") is None or item.get("amount") is None:
            item["needsReview"] = True
            item.setdefault("warnings", [])
            item["warnings"].append("필수 정보 누락")

    total = receipt.get("totalAmount")
    if total is not None and items:
        items_total = sum(item.get("amount") or 0 for item in items)
        if abs(items_total - total) > 1:
            receipt["needsReview"] = True
            receipt.setdefault("warnings", [])
            receipt["warnings"].append(f"합계 불일치: 품목합계 {items_total} ≠ 영수증합계 {total}")

    if receipt.get("isHandwritten"):
        receipt["needsReview"] = True
        for item in items:
            item["needsReview"] = True

    if receipt.get("confidence", 1.0) < 0.8:
        receipt["needsReview"] = True


async def _analyze_image(image_path: Path) -> Optional[dict]:
    """GPT-4o Vision으로 장끼/영수증 분석"""
    api_key = settings.OPENAI_API_KEY
    if not api_key:
        raise HTTPException(status_code=500, detail="OPENAI_API_KEY가 설정되지 않았습니다.")

    try:
        from openai import AsyncOpenAI
        client = AsyncOpenAI(api_key=api_key)

        with open(image_path, "rb") as f:
            b64 = base64.b64encode(f.read()).decode()
        mime = _get_mime(image_path.name)

        response = await client.chat.completions.create(
            model="gpt-4o",
            messages=[{
                "role": "user",
                "content": [
                    {"type": "text", "text": SYSTEM_PROMPT},
                    {"type": "image_url", "image_url": {
                        "url": f"data:{mime};base64,{b64}",
                        "detail": "high"
                    }},
                ],
            }],
            temperature=0.1,
            max_tokens=4096,
        )

        raw = response.choices[0].message.content or ""
        cleaned = raw.replace("```json", "").replace("```", "").strip()
        parsed = json.loads(cleaned)

        receipt = parsed.get("receipt", {})
        items = parsed.get("items", [])
        _validate_amounts(receipt, items)

        return parsed
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=422, detail=f"AI 응답 파싱 실패: {e}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI 분석 오류: {e}")


# ─────────────────────────────────────
# Pydantic Models
# ─────────────────────────────────────

class ReceiptItemUpdate(BaseModel):
    item_name: Optional[str] = None
    color: Optional[str] = None
    size: Optional[str] = None
    option_text: Optional[str] = None
    unit_price: Optional[float] = None
    quantity: Optional[float] = None
    amount: Optional[float] = None


class ReceiptUpdate(BaseModel):
    store_name: Optional[str] = None
    order_date: Optional[str] = None
    phone: Optional[str] = None
    total_amount: Optional[float] = None
    bank_info: Optional[str] = None
    memo: Optional[str] = None


# ─────────────────────────────────────
# 엔드포인트
# ─────────────────────────────────────

@router.post("/upload")
async def upload_and_analyze(
    token: str,
    file: UploadFile = File(...),
):
    """장끼/영수증 이미지 업로드 → GPT-4o 분석 → DB 저장"""
    ensure_receipt_tables()
    user = get_user_from_token(token)

    # 파일 저장
    ext = Path(file.filename or "receipt.jpg").suffix.lower() or ".jpg"
    filename = f"{uuid.uuid4()}{ext}"
    file_path = UPLOAD_DIR / filename

    content = await file.read()
    with open(file_path, "wb") as f:
        f.write(content)

    # AI 분석
    analysis = await _analyze_image(file_path)

    receipt_data = analysis.get("receipt", {})
    items_data = analysis.get("items", [])

    receipt_id = str(uuid.uuid4())

    with get_connection() as con:
        con.execute("""
            INSERT INTO receipts
            (id, user_id, image_filename, store_name, receipt_type, receipt_no,
             order_date, phone, total_amount, paid_amount, balance_amount,
             bank_info, memo, raw_text, image_quality, is_handwritten,
             confidence, needs_review, warnings)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            receipt_id,
            user["user_id"],
            filename,
            receipt_data.get("storeName"),
            receipt_data.get("receiptType", "장끼"),
            receipt_data.get("receiptNo"),
            receipt_data.get("orderDate"),
            receipt_data.get("phone"),
            receipt_data.get("totalAmount"),
            receipt_data.get("paidAmount"),
            receipt_data.get("balanceAmount"),
            receipt_data.get("bankInfo"),
            receipt_data.get("memo"),
            analysis.get("rawText"),
            receipt_data.get("imageQuality", "good"),
            1 if receipt_data.get("isHandwritten") else 0,
            receipt_data.get("confidence", 1.0),
            1 if receipt_data.get("needsReview") else 0,
            json.dumps(receipt_data.get("warnings", []), ensure_ascii=False),
        ))

        for item in items_data:
            item_id = str(uuid.uuid4())
            con.execute("""
                INSERT INTO receipt_items
                (id, receipt_id, line_no, raw_text, item_name, color, size,
                 option_text, unit_price, quantity, amount,
                 confidence, needs_review, warnings)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                item_id, receipt_id,
                item.get("lineNo"),
                item.get("rawText"),
                item.get("itemName"),
                item.get("color"),
                item.get("size"),
                item.get("optionText"),
                item.get("unitPrice"),
                item.get("quantity"),
                item.get("amount"),
                item.get("confidence", 1.0),
                1 if item.get("needsReview") else 0,
                json.dumps(item.get("warnings", []), ensure_ascii=False),
            ))

        con.commit()

    return {"receipt_id": receipt_id, "needs_review": bool(receipt_data.get("needsReview"))}


@router.get("/list")
def list_receipts(token: str, year: Optional[int] = None, month: Optional[int] = None):
    """영수증 목록 조회"""
    ensure_receipt_tables()
    user = get_user_from_token(token)

    with get_connection() as con:
        query = "SELECT * FROM receipts WHERE user_id = ?"
        params: list = [user["user_id"]]
        if year:
            query += " AND strftime('%Y', created_at) = ?"
            params.append(str(year))
        if month:
            query += " AND strftime('%m', created_at) = ?"
            params.append(f"{month:02d}")
        query += " ORDER BY created_at DESC"

        rows = con.execute(query, params).fetchall()
        cols = [d[0] for d in con.execute(query, params).description] if False else [
            "id", "user_id", "image_filename", "store_name", "receipt_type", "receipt_no",
            "order_date", "phone", "total_amount", "paid_amount", "balance_amount",
            "bank_info", "memo", "raw_text", "image_quality", "is_handwritten",
            "confidence", "needs_review", "warnings", "created_at"
        ]

    result = []
    for row in rows:
        d = dict(zip(cols, row))
        d["is_handwritten"] = bool(d["is_handwritten"])
        d["needs_review"] = bool(d["needs_review"])
        try:
            d["warnings"] = json.loads(d["warnings"] or "[]")
        except Exception:
            d["warnings"] = []
        result.append(d)
    return result


@router.get("/{receipt_id}")
def get_receipt(receipt_id: str, token: str):
    """영수증 단건 + 품목 조회"""
    ensure_receipt_tables()
    user = get_user_from_token(token)

    with get_connection() as con:
        row = con.execute(
            "SELECT * FROM receipts WHERE id = ? AND user_id = ?",
            (receipt_id, user["user_id"])
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="영수증을 찾을 수 없습니다.")

        cols = [
            "id", "user_id", "image_filename", "store_name", "receipt_type", "receipt_no",
            "order_date", "phone", "total_amount", "paid_amount", "balance_amount",
            "bank_info", "memo", "raw_text", "image_quality", "is_handwritten",
            "confidence", "needs_review", "warnings", "created_at"
        ]
        receipt = dict(zip(cols, row))
        receipt["is_handwritten"] = bool(receipt["is_handwritten"])
        receipt["needs_review"] = bool(receipt["needs_review"])
        try:
            receipt["warnings"] = json.loads(receipt["warnings"] or "[]")
        except Exception:
            receipt["warnings"] = []

        item_rows = con.execute(
            "SELECT * FROM receipt_items WHERE receipt_id = ? ORDER BY line_no",
            (receipt_id,)
        ).fetchall()
        item_cols = [
            "id", "receipt_id", "line_no", "raw_text", "item_name", "color", "size",
            "option_text", "unit_price", "quantity", "amount",
            "confidence", "needs_review", "warnings"
        ]
        items = []
        for ir in item_rows:
            d = dict(zip(item_cols, ir))
            d["needs_review"] = bool(d["needs_review"])
            try:
                d["warnings"] = json.loads(d["warnings"] or "[]")
            except Exception:
                d["warnings"] = []
            items.append(d)

        receipt["items"] = items
    return receipt


@router.put("/{receipt_id}")
def update_receipt(receipt_id: str, token: str, data: ReceiptUpdate):
    """영수증 헤더 수정"""
    ensure_receipt_tables()
    user = get_user_from_token(token)

    fields = {k: v for k, v in data.dict().items() if v is not None}
    if not fields:
        return {"success": True}

    set_clause = ", ".join(f"{k} = ?" for k in fields)
    with get_connection() as con:
        con.execute(
            f"UPDATE receipts SET {set_clause} WHERE id = ? AND user_id = ?",
            list(fields.values()) + [receipt_id, user["user_id"]]
        )
        con.commit()
    return {"success": True}


@router.put("/items/{item_id}")
def update_receipt_item(item_id: str, token: str, data: ReceiptItemUpdate):
    """품목 수정"""
    ensure_receipt_tables()
    get_user_from_token(token)

    fields = {k: v for k, v in data.dict().items() if v is not None}
    if not fields:
        return {"success": True}

    col_map = {
        "item_name": "item_name", "color": "color", "size": "size",
        "option_text": "option_text", "unit_price": "unit_price",
        "quantity": "quantity", "amount": "amount",
    }
    db_fields = {col_map[k]: v for k, v in fields.items() if k in col_map}

    # 금액 재검산
    if "unit_price" in db_fields or "quantity" in db_fields or "amount" in db_fields:
        with get_connection() as con:
            row = con.execute(
                "SELECT unit_price, quantity, amount FROM receipt_items WHERE id = ?",
                (item_id,)
            ).fetchone()
        if row:
            u = db_fields.get("unit_price", row[0])
            q = db_fields.get("quantity", row[1])
            a = db_fields.get("amount", row[2])
            if u is not None and q is not None and a is not None:
                calc = u * q
                nr = 1 if abs(calc - a) > 1 else 0
                db_fields["needs_review"] = nr

    set_clause = ", ".join(f"{k} = ?" for k in db_fields)
    with get_connection() as con:
        con.execute(
            f"UPDATE receipt_items SET {set_clause} WHERE id = ?",
            list(db_fields.values()) + [item_id]
        )
        con.commit()
    return {"success": True}


@router.delete("/{receipt_id}")
def delete_receipt(receipt_id: str, token: str):
    """영수증 삭제"""
    ensure_receipt_tables()
    user = get_user_from_token(token)

    with get_connection() as con:
        row = con.execute(
            "SELECT image_filename FROM receipts WHERE id = ? AND user_id = ?",
            (receipt_id, user["user_id"])
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="영수증을 찾을 수 없습니다.")

        # 이미지 파일 삭제
        if row[0]:
            img_path = UPLOAD_DIR / row[0]
            if img_path.exists():
                img_path.unlink()

        con.execute("DELETE FROM receipt_items WHERE receipt_id = ?", (receipt_id,))
        con.execute("DELETE FROM receipts WHERE id = ?", (receipt_id,))
        con.commit()
    return {"success": True}


@router.get("/{receipt_id}/excel")
def download_excel(receipt_id: str, token: str):
    """영수증 엑셀 다운로드"""
    ensure_receipt_tables()
    user = get_user_from_token(token)

    receipt_full = get_receipt(receipt_id, token)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "영수증정리"

        # 헤더
        headers = ["거래일", "거래처명", "영수증번호", "품명", "색상", "사이즈", "옵션",
                   "단가", "수량", "금액", "합계", "연락처", "계좌정보", "확인필요", "비고"]
        header_fill = PatternFill("solid", fgColor="366092")
        header_font = Font(color="FFFFFF", bold=True)
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # 데이터 행
        review_fill = PatternFill("solid", fgColor="FFF2CC")
        for item in receipt_full["items"]:
            row_data = [
                receipt_full.get("order_date", ""),
                receipt_full.get("store_name", ""),
                receipt_full.get("receipt_no", ""),
                item.get("item_name", ""),
                item.get("color", ""),
                item.get("size", ""),
                item.get("option_text", ""),
                item.get("unit_price"),
                item.get("quantity"),
                item.get("amount"),
                receipt_full.get("total_amount"),
                receipt_full.get("phone", ""),
                receipt_full.get("bank_info", ""),
                "O" if item.get("needs_review") else "",
                ", ".join(item.get("warnings", [])),
            ]
            ws.append(row_data)
            if item.get("needs_review"):
                for ci in range(1, len(headers) + 1):
                    ws.cell(row=ws.max_row, column=ci).fill = review_fill

        # 컬럼 너비
        col_widths = [12, 20, 15, 28, 10, 8, 18, 10, 8, 12, 12, 15, 30, 10, 30]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        store = receipt_full.get("store_name") or "영수증"
        dt = (receipt_full.get("order_date") or datetime.now().strftime("%Y%m%d")).replace("-", "")
        filename = f"영수증정리_{store}_{dt}.xlsx"

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl이 설치되지 않았습니다.")


@router.get("/image/{filename}")
def get_image(filename: str, token: str):
    """영수증 이미지 반환"""
    get_user_from_token(token)
    img_path = UPLOAD_DIR / filename
    if not img_path.exists():
        raise HTTPException(status_code=404, detail="이미지를 찾을 수 없습니다.")

    from fastapi.responses import FileResponse
    return FileResponse(img_path)

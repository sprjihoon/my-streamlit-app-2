"""
backend/app/api/invoices.py - 인보이스 목록 및 관리 API
"""
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from typing import List, Optional, Dict, Any
import pandas as pd
import io

from logic.db import get_connection
from backend.app.api.logs import add_log

router = APIRouter(prefix="/invoices", tags=["invoices"])


# ─────────────────────────────────────
# 권한 체크 헬퍼
# ─────────────────────────────────────

def check_admin(token: Optional[str]) -> tuple:
    """관리자 권한 확인, (is_admin, nickname) 반환"""
    if not token:
        return False, None
    with get_connection() as con:
        result = con.execute(
            "SELECT u.is_admin, u.nickname FROM sessions s JOIN users u ON s.user_id = u.user_id WHERE s.token = ?",
            (token,)
        ).fetchone()
        if result:
            return bool(result[0]), result[1]
    return False, None


def get_user_nickname(token: Optional[str]) -> str:
    """토큰에서 닉네임 가져오기"""
    if not token:
        return '시스템'
    with get_connection() as con:
        result = con.execute(
            "SELECT u.nickname FROM sessions s JOIN users u ON s.user_id = u.user_id WHERE s.token = ?",
            (token,)
        ).fetchone()
        return result[0] if result else '시스템'


# ─────────────────────────────────────
# API Endpoints
# ─────────────────────────────────────

@router.get("")
@router.get("/")
async def list_invoices(
    period: Optional[str] = None,
    vendor: Optional[str] = None,
    status: Optional[str] = None
):
    """인보이스 목록 조회"""
    try:
        with get_connection() as con:
            # 테이블 존재 확인
            tables = [row[0] for row in con.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()]
            
            if "invoices" not in tables:
                return {"invoices": [], "total": 0, "sum_amount": 0}
            
            # 컬럼 존재 확인
            cols = [c[1] for c in con.execute("PRAGMA table_info(invoices);")]
            has_modified_by = 'modified_by' in cols
            has_confirmed_by = 'confirmed_by' in cols
            
            # 기본 쿼리
            select_cols = """
                    i.invoice_id,
                    i.vendor_id,
                    COALESCE(v.name, v.vendor, i.vendor_id) as vendor_name,
                    i.period_from,
                    i.period_to,
                    i.total_amount,
                    COALESCE(i.status, '미확정') as status,
                    i.created_at"""
            
            if has_modified_by:
                select_cols += ", i.modified_by, i.modified_at"
            if has_confirmed_by:
                select_cols += ", i.confirmed_by, i.confirmed_at"
            
            query = f"""
                SELECT {select_cols}
                FROM invoices i
                LEFT JOIN vendors v ON i.vendor_id = v.vendor_id
                WHERE 1=1
            """
            params = []
            
            # 필터 적용
            if period:
                query += " AND strftime('%Y-%m', i.period_from) = ?"
                params.append(period)
            
            if vendor:
                query += " AND (v.vendor = ? OR v.name = ?)"
                params.extend([vendor, vendor])
            
            if status:
                query += " AND i.status = ?"
                params.append(status)
            
            query += " ORDER BY i.invoice_id DESC"
            
            df = pd.read_sql(query, con, params=params)
            
            # 합계 계산
            df['total_amount'] = pd.to_numeric(df['total_amount'], errors='coerce').fillna(0)
            sum_amount = int(df['total_amount'].sum())
            
            invoices = []
            for _, row in df.iterrows():
                inv_data = {
                    "invoice_id": int(row['invoice_id']),
                    "vendor_id": row['vendor_id'],
                    "vendor": str(row['vendor_name']) if row['vendor_name'] else '',
                    "period_from": str(row['period_from']) if row['period_from'] else '',
                    "period_to": str(row['period_to']) if row['period_to'] else '',
                    "total_amount": int(row['total_amount']),
                    "status": str(row['status']),
                    "created_at": str(row['created_at']) if row['created_at'] else '',
                    "modified_by": str(row['modified_by']) if has_modified_by and pd.notna(row.get('modified_by')) else None,
                    "modified_at": str(row['modified_at']) if has_modified_by and pd.notna(row.get('modified_at')) else None,
                    "confirmed_by": str(row['confirmed_by']) if has_confirmed_by and pd.notna(row.get('confirmed_by')) else None,
                    "confirmed_at": str(row['confirmed_at']) if has_confirmed_by and pd.notna(row.get('confirmed_at')) else None,
                }
                invoices.append(inv_data)
            
            # 사용 가능한 기간 목록
            periods_df = pd.read_sql(
                "SELECT DISTINCT strftime('%Y-%m', period_from) as ym FROM invoices ORDER BY ym DESC",
                con
            )
            periods = periods_df['ym'].dropna().tolist()
            
            return {
                "invoices": invoices,
                "total": len(invoices),
                "sum_amount": sum_amount,
                "periods": periods
            }
    
    except Exception as e:
        return {"invoices": [], "total": 0, "sum_amount": 0, "error": str(e)}


@router.get("/{invoice_id}/export/pdf")
async def export_single_invoice_pdf(invoice_id: int):
    """단일 인보이스 PDF 다운로드 (물류대행 서비스 대금청구서 양식)"""
    try:
        from logic.invoice_pdf_v2 import create_billing_invoice_pdf
        from datetime import datetime
        
        with get_connection() as con:
            # 인보이스 컬럼 확인 및 마이그레이션
            cols = [c[1] for c in con.execute("PRAGMA table_info(invoices);")]
            if 'confirmed_by' not in cols:
                try:
                    con.execute("ALTER TABLE invoices ADD COLUMN confirmed_by TEXT")
                    con.commit()
                except:
                    pass
            has_confirmed_by = True  # 컬럼이 있거나 추가됨
            
            # 인보이스 정보 (확정자 닉네임 포함)
            select_cols = """
                i.invoice_id,
                i.vendor_id,
                COALESCE(v.name, v.vendor) as vendor_name,
                i.period_from,
                i.period_to,
                i.total_amount,
                i.created_at
            """
            if has_confirmed_by:
                select_cols += ", i.confirmed_by"
            
            inv_df = pd.read_sql(
                f"""
                SELECT {select_cols}
                FROM invoices i
                LEFT JOIN vendors v ON i.vendor_id = v.vendor_id
                WHERE i.invoice_id = ?
                """,
                con, params=[invoice_id]
            )
            
            if inv_df.empty:
                raise HTTPException(status_code=404, detail="Invoice not found")
            
            inv = inv_df.iloc[0]
            vendor_name = str(inv['vendor_name']) if inv['vendor_name'] else 'Unknown'
            period_from = str(inv['period_from']) if inv['period_from'] else ''
            period_to = str(inv['period_to']) if inv['period_to'] else ''
            
            # 담당자 = 인보이스 확정자 닉네임
            confirmed_by = ""
            if has_confirmed_by and pd.notna(inv.get('confirmed_by')):
                confirmed_by = str(inv['confirmed_by'])
            
            # 항목 조회
            items_df = pd.read_sql(
                "SELECT item_name as 항목, qty as 수량, unit_price as 단가, amount as 금액, remark as 비고 FROM invoice_items WHERE invoice_id = ?",
                con, params=[invoice_id]
            )
            
            items = items_df.to_dict('records')
        
        # 청구일자
        invoice_date = datetime.now().strftime("%Y-%m-%d")
        
        # 건명 생성
        period_str = period_from[:7].replace("-", "년 ") + "월" if period_from else ""
        title = f"{period_str} 풀필먼트 서비스 대금"
        
        # 수신자 - "업체별칭 대표님 귀하"
        recipient_name = f"{vendor_name} 대표님 귀하"
        
        # 회사 설정 조회 (DB에서) - 테이블 및 기본값 보장
        with get_connection() as con2:
            # 테이블 존재 확인 및 생성
            table_exists = con2.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name='company_settings'"
            ).fetchone()
            
            if not table_exists:
                con2.execute("""
                    CREATE TABLE company_settings(
                        id              INTEGER PRIMARY KEY CHECK (id = 1),
                        company_name    TEXT DEFAULT '회사명',
                        business_number TEXT DEFAULT '000-00-00000',
                        address         TEXT DEFAULT '주소를 입력하세요',
                        business_type   TEXT DEFAULT '서비스',
                        business_item   TEXT DEFAULT '물류대행',
                        bank_name       TEXT DEFAULT '은행명',
                        account_holder  TEXT DEFAULT '예금주',
                        account_number  TEXT DEFAULT '계좌번호',
                        representative  TEXT DEFAULT '대표자명',
                        updated_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                con2.commit()
            
            # 기본 레코드 삽입 (없으면)
            row_check = con2.execute("SELECT 1 FROM company_settings WHERE id = 1").fetchone()
            if not row_check:
                con2.execute("""
                    INSERT INTO company_settings (id, company_name, business_number, address, 
                        business_type, business_item, bank_name, account_holder, account_number, representative)
                    VALUES (1, '틸리언', '766-55-00323', '대구시 동구 첨단로8길 8 씨제이빌딩302호',
                        '서비스', '포장 및 충전업', '카카오뱅크', '장지훈', '3333-02-9946468', '장지훈')
                """)
                con2.commit()
            
            company_row = con2.execute("""
                SELECT company_name, business_number, address, business_type, business_item,
                       bank_name, account_holder, account_number, representative
                FROM company_settings WHERE id = 1
            """).fetchone()
        
        if company_row:
            supplier_info = {
                "사업자번호": company_row[1] or "",
                "상호": company_row[0] or "",
                "소재지": company_row[2] or "",
                "업태": company_row[3] or "",
                "종목": company_row[4] or "",
            }
            bank_info = {
                "은행명": company_row[5] or "",
                "예금주": company_row[6] or "",
                "계좌번호": company_row[7] or "",
            }
            representative = company_row[8] or ""
            company_display_name = company_row[0] or ""
        else:
            # 기본값 (설정이 없는 경우)
            supplier_info = {
                "사업자번호": "",
                "상호": "",
                "소재지": "",
                "업태": "",
                "종목": "",
            }
            bank_info = {
                "은행명": "",
                "예금주": "",
                "계좌번호": "",
            }
            representative = ""
            company_display_name = ""
        
        # 지급기한 - 청구 기간 시작월 + 1개월의 5일
        # period_from이 "2025-11-01"이면 -> "2025년 12월 05일"
        if period_from:
            try:
                from dateutil.relativedelta import relativedelta
                period_dt = datetime.strptime(period_from[:10], "%Y-%m-%d")
                next_month = period_dt + relativedelta(months=1)
                payment_deadline = f"{next_month.year}년 {next_month.month:02d}월 05일"
            except:
                payment_deadline = ""
        else:
            payment_deadline = ""
        
        # PDF 생성
        pdf_bytes = create_billing_invoice_pdf(
            invoice_id=invoice_id,
            invoice_date=invoice_date,
            recipient_name=recipient_name,
            title=title,
            supplier_info=supplier_info,
            items=items,
            payment_deadline=payment_deadline,
            bank_info=bank_info,
            stamp_holder=representative,  # 대표 - DB에서 가져옴
            manager=confirmed_by,         # 담당 - 인보이스 확정자 닉네임
            company_name=company_display_name,  # 하단 회사명 - DB에서 가져옴
        )
        
        # 파일명 생성
        filename = f"invoice_{invoice_id}_{period_from[:7] if period_from else 'unknown'}.pdf"
        
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{invoice_id}/export/xlsx")
async def export_single_invoice_xlsx(invoice_id: int):
    """단일 인보이스 엑셀 다운로드 (PDF와 동일한 양식)"""
    try:
        from datetime import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        import re
        
        with get_connection() as con:
            # 인보이스 컬럼 확인
            cols = [c[1] for c in con.execute("PRAGMA table_info(invoices);")]
            has_confirmed_by = 'confirmed_by' in cols
            
            # 인보이스 정보
            select_cols = """
                i.invoice_id,
                i.vendor_id,
                COALESCE(v.name, v.vendor) as vendor_name,
                i.period_from,
                i.period_to,
                i.total_amount,
                i.created_at
            """
            if has_confirmed_by:
                select_cols += ", i.confirmed_by"
            
            inv_df = pd.read_sql(
                f"""
                SELECT {select_cols}
                FROM invoices i
                LEFT JOIN vendors v ON i.vendor_id = v.vendor_id
                WHERE i.invoice_id = ?
                """,
                con, params=[invoice_id]
            )
            
            if inv_df.empty:
                raise HTTPException(status_code=404, detail="Invoice not found")
            
            inv = inv_df.iloc[0]
            vendor_name = str(inv['vendor_name']) if inv['vendor_name'] else 'Unknown'
            period_from = str(inv['period_from']) if inv['period_from'] else ''
            period_to = str(inv['period_to']) if inv['period_to'] else ''
            
            # 담당자 = 인보이스 확정자 닉네임
            confirmed_by = ""
            if has_confirmed_by and pd.notna(inv.get('confirmed_by')):
                confirmed_by = str(inv['confirmed_by'])
            
            # 항목 조회
            items_df = pd.read_sql(
                "SELECT item_name as 항목, qty as 수량, unit_price as 단가, amount as 금액, remark as 비고 FROM invoice_items WHERE invoice_id = ?",
                con, params=[invoice_id]
            )
            
            # 회사 설정 조회
            company_row = con.execute("""
                SELECT company_name, business_number, address, business_type, business_item,
                       bank_name, account_holder, account_number, representative
                FROM company_settings WHERE id = 1
            """).fetchone()
        
        # 회사 정보 설정
        if company_row:
            company_name = company_row[0] or ""
            business_number = company_row[1] or ""
            address = company_row[2] or ""
            business_type = company_row[3] or ""
            business_item = company_row[4] or ""
            bank_name = company_row[5] or ""
            account_holder = company_row[6] or ""
            account_number = company_row[7] or ""
            representative = company_row[8] or ""
        else:
            company_name = business_number = address = business_type = business_item = ""
            bank_name = account_holder = account_number = representative = ""
        
        # 청구일자
        invoice_date = datetime.now().strftime("%Y-%m-%d")
        
        # 건명 생성
        period_str = period_from[:7].replace("-", "년 ") + "월" if period_from else ""
        title = f"{period_str} 풀필먼트 서비스 대금"
        
        # 수신자
        recipient_name = f"{vendor_name} 대표님 귀하"
        
        # 문서번호
        doc_number = f"{invoice_id:05d}-{invoice_date.replace('-', '')[:6]}"
        
        # 지급기한
        if period_from:
            try:
                from dateutil.relativedelta import relativedelta
                period_dt = datetime.strptime(period_from[:10], "%Y-%m-%d")
                next_month = period_dt + relativedelta(months=1)
                payment_deadline = f"{next_month.year}년 {next_month.month:02d}월 05일"
            except:
                payment_deadline = ""
        else:
            payment_deadline = ""
        
        # 엑셀 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "청구서"
        
        # 스타일 정의
        title_font = Font(name='맑은 고딕', size=18, bold=True)
        header_font = Font(name='맑은 고딕', size=10, bold=True)
        body_font = Font(name='맑은 고딕', size=9)
        small_font = Font(name='맑은 고딕', size=8)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        gray_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        light_gray_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        current_row = 1
        
        # 1. 제목
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = "물류대행 서비스 대금청구서"
        ws[f'A{current_row}'].font = title_font
        ws[f'A{current_row}'].alignment = center_align
        current_row += 2
        
        # 2. 문서번호/청구일자
        ws[f'A{current_row}'] = "문서번호"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].fill = gray_fill
        ws[f'A{current_row}'].border = thin_border
        ws.merge_cells(f'B{current_row}:C{current_row}')
        ws[f'B{current_row}'] = doc_number
        ws[f'B{current_row}'].font = body_font
        ws[f'B{current_row}'].border = thin_border
        ws[f'D{current_row}'] = "청구일자"
        ws[f'D{current_row}'].font = header_font
        ws[f'D{current_row}'].fill = gray_fill
        ws[f'D{current_row}'].border = thin_border
        ws.merge_cells(f'E{current_row}:F{current_row}')
        ws[f'E{current_row}'] = invoice_date
        ws[f'E{current_row}'].font = body_font
        ws[f'E{current_row}'].border = thin_border
        current_row += 1
        
        # 3. 수신/건명
        ws[f'A{current_row}'] = "수신"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].fill = gray_fill
        ws[f'A{current_row}'].border = thin_border
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = recipient_name
        ws[f'B{current_row}'].font = body_font
        ws[f'B{current_row}'].border = thin_border
        current_row += 1
        
        ws[f'A{current_row}'] = "건명"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].fill = gray_fill
        ws[f'A{current_row}'].border = thin_border
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = title
        ws[f'B{current_row}'].font = body_font
        ws[f'B{current_row}'].border = thin_border
        current_row += 1
        
        # 4. 공급자 정보
        ws[f'A{current_row}'] = "공급자"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].fill = gray_fill
        ws[f'A{current_row}'].border = thin_border
        ws.merge_cells(f'A{current_row}:A{current_row+2}')
        
        ws[f'B{current_row}'] = "사업자번호"
        ws[f'B{current_row}'].font = small_font
        ws[f'B{current_row}'].fill = light_gray_fill
        ws[f'B{current_row}'].border = thin_border
        ws[f'C{current_row}'] = business_number
        ws[f'C{current_row}'].font = small_font
        ws[f'C{current_row}'].border = thin_border
        ws[f'D{current_row}'] = "상호"
        ws[f'D{current_row}'].font = small_font
        ws[f'D{current_row}'].fill = light_gray_fill
        ws[f'D{current_row}'].border = thin_border
        ws.merge_cells(f'E{current_row}:F{current_row}')
        ws[f'E{current_row}'] = company_name
        ws[f'E{current_row}'].font = small_font
        ws[f'E{current_row}'].border = thin_border
        current_row += 1
        
        ws[f'B{current_row}'] = "소재지"
        ws[f'B{current_row}'].font = small_font
        ws[f'B{current_row}'].fill = light_gray_fill
        ws[f'B{current_row}'].border = thin_border
        ws.merge_cells(f'C{current_row}:F{current_row}')
        ws[f'C{current_row}'] = address
        ws[f'C{current_row}'].font = small_font
        ws[f'C{current_row}'].border = thin_border
        current_row += 1
        
        ws[f'B{current_row}'] = "업태"
        ws[f'B{current_row}'].font = small_font
        ws[f'B{current_row}'].fill = light_gray_fill
        ws[f'B{current_row}'].border = thin_border
        ws[f'C{current_row}'] = business_type
        ws[f'C{current_row}'].font = small_font
        ws[f'C{current_row}'].border = thin_border
        ws[f'D{current_row}'] = "종목"
        ws[f'D{current_row}'].font = small_font
        ws[f'D{current_row}'].fill = light_gray_fill
        ws[f'D{current_row}'].border = thin_border
        ws.merge_cells(f'E{current_row}:F{current_row}')
        ws[f'E{current_row}'] = business_item
        ws[f'E{current_row}'].font = small_font
        ws[f'E{current_row}'].border = thin_border
        current_row += 2
        
        # 5. 항목 테이블 헤더
        headers = ["No", "품명", "수량", "단가", "금액", "비고"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = gray_fill
            cell.border = thin_border
            cell.alignment = center_align
        current_row += 1
        
        # 항목 데이터
        subtotal = 0
        for idx, (_, row) in enumerate(items_df.iterrows(), 1):
            qty = int(row['수량']) if pd.notna(row['수량']) else 0
            unit_price = int(row['단가']) if pd.notna(row['단가']) else 0
            amount = int(row['금액']) if pd.notna(row['금액']) else qty * unit_price
            subtotal += amount
            
            ws.cell(row=current_row, column=1, value=idx).border = thin_border
            ws.cell(row=current_row, column=1).alignment = center_align
            ws.cell(row=current_row, column=2, value=str(row['항목'])).border = thin_border
            ws.cell(row=current_row, column=3, value=f"{qty:,}" if qty else "").border = thin_border
            ws.cell(row=current_row, column=3).alignment = right_align
            ws.cell(row=current_row, column=4, value=f"{unit_price:,}" if unit_price else "").border = thin_border
            ws.cell(row=current_row, column=4).alignment = right_align
            ws.cell(row=current_row, column=5, value=f"{amount:,}" if amount else "").border = thin_border
            ws.cell(row=current_row, column=5).alignment = right_align
            ws.cell(row=current_row, column=6, value=str(row['비고']) if pd.notna(row['비고']) else "").border = thin_border
            
            for col in range(1, 7):
                ws.cell(row=current_row, column=col).font = body_font
            current_row += 1
        
        current_row += 1
        
        # 6. 합계
        vat = int(subtotal * 0.1)
        total = subtotal + vat
        
        ws[f'A{current_row}'] = "합계 금액"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].fill = gray_fill
        ws[f'A{current_row}'].border = thin_border
        ws[f'B{current_row}'] = f"₩ {subtotal:,}"
        ws[f'B{current_row}'].font = body_font
        ws[f'B{current_row}'].border = thin_border
        ws[f'B{current_row}'].alignment = right_align
        ws[f'C{current_row}'] = "부가세"
        ws[f'C{current_row}'].font = header_font
        ws[f'C{current_row}'].fill = gray_fill
        ws[f'C{current_row}'].border = thin_border
        ws[f'D{current_row}'] = f"₩ {vat:,}"
        ws[f'D{current_row}'].font = body_font
        ws[f'D{current_row}'].border = thin_border
        ws[f'D{current_row}'].alignment = right_align
        ws[f'E{current_row}'] = "청구금액"
        ws[f'E{current_row}'].font = header_font
        ws[f'E{current_row}'].fill = gray_fill
        ws[f'E{current_row}'].border = thin_border
        ws[f'F{current_row}'] = f"₩ {total:,}"
        ws[f'F{current_row}'].font = Font(name='맑은 고딕', size=11, bold=True)
        ws[f'F{current_row}'].border = thin_border
        ws[f'F{current_row}'].alignment = right_align
        current_row += 2
        
        # 7. 지급기한/계좌정보
        ws[f'A{current_row}'] = "지급기한"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].fill = gray_fill
        ws[f'A{current_row}'].border = thin_border
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = payment_deadline
        ws[f'B{current_row}'].font = body_font
        ws[f'B{current_row}'].border = thin_border
        current_row += 1
        
        ws[f'A{current_row}'] = "계좌정보"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].fill = gray_fill
        ws[f'A{current_row}'].border = thin_border
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = f"{bank_name}  {account_number}  {account_holder}"
        ws[f'B{current_row}'].font = body_font
        ws[f'B{current_row}'].border = thin_border
        current_row += 3
        
        # 8. 하단 - 위와 같이 청구합니다
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = "위와 같이 청구합니다."
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].alignment = center_align
        current_row += 2
        
        # 날짜 (한국어 형식)
        try:
            dt = datetime.strptime(invoice_date, "%Y-%m-%d")
            weekdays = ['월', '화', '수', '목', '금', '토', '일']
            date_str = f"{dt.year}년 {dt.month:02d}월 {dt.day:02d}일 {weekdays[dt.weekday()]}요일"
        except:
            date_str = invoice_date
        
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = date_str
        ws[f'A{current_row}'].font = body_font
        ws[f'A{current_row}'].alignment = center_align
        current_row += 2
        
        # 회사명
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = company_name
        ws[f'A{current_row}'].font = title_font
        ws[f'A{current_row}'].alignment = center_align
        current_row += 1
        
        # 담당자/대표자 정보
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = f"담당: {confirmed_by or '-'}  /  대표: {representative or '-'}"
        ws[f'A{current_row}'].font = small_font
        ws[f'A{current_row}'].alignment = center_align
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 25
        
        # 엑셀 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        period = period_from[:7] if period_from else ''
        ascii_filename = f"invoice_{invoice_id}_{period}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={ascii_filename}"
            }
        )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{invoice_id}")
async def get_invoice_detail(invoice_id: int):
    """인보이스 상세 조회"""
    try:
        with get_connection() as con:
            # 인보이스 기본 정보
            inv_df = pd.read_sql(
                """
                SELECT 
                    i.invoice_id,
                    i.vendor_id,
                    COALESCE(v.name, v.vendor) as vendor_name,
                    i.period_from,
                    i.period_to,
                    i.total_amount,
                    COALESCE(i.status, '미확정') as status,
                    i.created_at
                FROM invoices i
                LEFT JOIN vendors v ON i.vendor_id = v.vendor_id
                WHERE i.invoice_id = ?
                """,
                con, params=[invoice_id]
            )
            
            if inv_df.empty:
                raise HTTPException(status_code=404, detail="Invoice not found")
            
            inv = inv_df.iloc[0]
            
            # 인보이스 항목
            items_df = pd.read_sql(
                "SELECT item_name, qty, unit_price, amount, remark FROM invoice_items WHERE invoice_id = ?",
                con, params=[invoice_id]
            )
            
            items = []
            for _, row in items_df.iterrows():
                items.append({
                    "항목": str(row['item_name']) if row['item_name'] else '',
                    "수량": int(row['qty']) if pd.notna(row['qty']) else 0,
                    "단가": int(row['unit_price']) if pd.notna(row['unit_price']) else 0,
                    "금액": int(row['amount']) if pd.notna(row['amount']) else 0,
                    "비고": str(row['remark']) if row['remark'] else ''
                })
            
            return {
                "invoice_id": int(inv['invoice_id']),
                "vendor": str(inv['vendor_name']) if inv['vendor_name'] else '',
                "period_from": str(inv['period_from']) if inv['period_from'] else '',
                "period_to": str(inv['period_to']) if inv['period_to'] else '',
                "total_amount": int(inv['total_amount']) if pd.notna(inv['total_amount']) else 0,
                "status": str(inv['status']),
                "created_at": str(inv['created_at']) if inv['created_at'] else '',
                "items": items
            }
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/{invoice_id}")
async def delete_invoice(invoice_id: int, token: Optional[str] = None):
    """인보이스 삭제 (관리자만)"""
    # 관리자 권한 체크
    is_admin, nickname = check_admin(token)
    if not is_admin:
        raise HTTPException(status_code=403, detail="관리자 권한이 필요합니다.")
    
    try:
        with get_connection() as con:
            # 삭제 전 인보이스 정보 가져오기
            inv = con.execute(
                "SELECT vendor_id FROM invoices WHERE invoice_id = ?", (invoice_id,)
            ).fetchone()
            vendor_name = inv[0] if inv else "알 수 없음"
            
            con.execute("DELETE FROM invoice_items WHERE invoice_id = ?", (invoice_id,))
            con.execute("DELETE FROM invoices WHERE invoice_id = ?", (invoice_id,))
            con.commit()
        
        # 로그 기록
        add_log(
            action_type="인보이스 삭제",
            target_type="invoice",
            target_id=str(invoice_id),
            target_name=vendor_name,
            user_nickname=nickname,
            details=f"인보이스 ID {invoice_id} 삭제"
        )
        
        return {"status": "success", "deleted": invoice_id}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/{invoice_id}/confirm")
async def confirm_invoice(invoice_id: int, token: Optional[str] = None, user_nickname: Optional[str] = None):
    """인보이스 확정"""
    try:
        with get_connection() as con:
            # 컬럼 존재 확인 및 추가
            ensure_invoice_user_columns(con)
            
            # 사용자 닉네임 가져오기
            nickname = user_nickname or get_nickname_from_token(con, token) or '시스템'
            
            # 인보이스 정보 가져오기
            inv = con.execute(
                "SELECT vendor_id FROM invoices WHERE invoice_id = ?", (invoice_id,)
            ).fetchone()
            vendor_name = inv[0] if inv else "알 수 없음"
            
            con.execute(
                "UPDATE invoices SET status = '확정', confirmed_by = ?, confirmed_at = CURRENT_TIMESTAMP WHERE invoice_id = ?",
                (nickname, invoice_id)
            )
            con.commit()
        
        # 로그 기록
        add_log(
            action_type="인보이스 확정",
            target_type="invoice",
            target_id=str(invoice_id),
            target_name=vendor_name,
            user_nickname=nickname,
            details=f"인보이스 ID {invoice_id} 확정 처리"
        )
        
        return {"status": "success", "invoice_id": invoice_id, "new_status": "확정", "confirmed_by": nickname}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/{invoice_id}/unconfirm")
async def unconfirm_invoice(invoice_id: int, token: Optional[str] = None, user_nickname: Optional[str] = None):
    """인보이스 미확정으로 변경"""
    try:
        with get_connection() as con:
            # 컬럼 존재 확인 및 추가
            ensure_invoice_user_columns(con)
            
            # 사용자 닉네임 가져오기
            nickname = user_nickname or get_nickname_from_token(con, token) or '시스템'
            
            # 인보이스 정보 가져오기
            inv = con.execute(
                "SELECT vendor_id FROM invoices WHERE invoice_id = ?", (invoice_id,)
            ).fetchone()
            vendor_id = inv[0] if inv else "알 수 없음"
            
            con.execute(
                "UPDATE invoices SET status = '미확정', confirmed_by = NULL, confirmed_at = NULL, modified_by = ?, modified_at = CURRENT_TIMESTAMP WHERE invoice_id = ?",
                (nickname, invoice_id)
            )
            con.commit()
        
        # 로그 기록
        add_log(
            action_type="인보이스 미확정",
            target_type="invoice",
            target_id=str(invoice_id),
            target_name=vendor_id,
            user_nickname=nickname,
            details=f"인보이스 ID {invoice_id} 미확정 처리"
        )
        
        return {"status": "success", "invoice_id": invoice_id, "new_status": "미확정", "modified_by": nickname}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


from pydantic import BaseModel

class InvoiceItemUpdate(BaseModel):
    항목: str
    수량: int
    단가: int
    금액: int
    비고: str = ""

class InvoiceUpdateRequest(BaseModel):
    items: List[InvoiceItemUpdate]
    user_nickname: Optional[str] = None


def ensure_invoice_user_columns(con):
    """인보이스 테이블에 사용자 관련 컬럼 추가"""
    cols = [c[1] for c in con.execute("PRAGMA table_info(invoices);")]
    if 'modified_by' not in cols:
        con.execute("ALTER TABLE invoices ADD COLUMN modified_by TEXT;")
    if 'modified_at' not in cols:
        con.execute("ALTER TABLE invoices ADD COLUMN modified_at DATETIME;")
    if 'confirmed_by' not in cols:
        con.execute("ALTER TABLE invoices ADD COLUMN confirmed_by TEXT;")
    if 'confirmed_at' not in cols:
        con.execute("ALTER TABLE invoices ADD COLUMN confirmed_at DATETIME;")


def get_nickname_from_token(con, token: Optional[str]) -> Optional[str]:
    """토큰에서 사용자 닉네임 가져오기"""
    if not token:
        return None
    try:
        result = con.execute(
            "SELECT u.nickname FROM sessions s JOIN users u ON s.user_id = u.user_id WHERE s.token = ?",
            (token,)
        ).fetchone()
        return result[0] if result else None
    except:
        return None


@router.put("/{invoice_id}/items")
async def update_invoice_items(invoice_id: int, request: InvoiceUpdateRequest, token: Optional[str] = None):
    """인보이스 항목 수정"""
    try:
        with get_connection() as con:
            # 컬럼 존재 확인 및 추가
            ensure_invoice_user_columns(con)
            
            # 기존 인보이스 확인
            existing = con.execute(
                "SELECT invoice_id, vendor_id FROM invoices WHERE invoice_id = ?", (invoice_id,)
            ).fetchone()
            
            if not existing:
                raise HTTPException(status_code=404, detail="Invoice not found")
            
            vendor_name = existing[1] if existing[1] else "알 수 없음"
            
            # 사용자 닉네임
            nickname = request.user_nickname or get_nickname_from_token(con, token) or '시스템'
            
            # 기존 항목 삭제
            con.execute("DELETE FROM invoice_items WHERE invoice_id = ?", (invoice_id,))
            
            # 새 항목 삽입
            total_amount = 0
            for item in request.items:
                con.execute(
                    "INSERT INTO invoice_items (invoice_id, item_name, qty, unit_price, amount, remark) VALUES (?, ?, ?, ?, ?, ?)",
                    (invoice_id, item.항목, item.수량, item.단가, item.금액, item.비고)
                )
                total_amount += item.금액
            
            # 총액 및 수정자 업데이트
            con.execute(
                "UPDATE invoices SET total_amount = ?, modified_by = ?, modified_at = CURRENT_TIMESTAMP WHERE invoice_id = ?",
                (total_amount, nickname, invoice_id)
            )
            
            con.commit()
        
        # 로그 기록
        add_log(
            action_type="인보이스 수정",
            target_type="invoice",
            target_id=str(invoice_id),
            target_name=vendor_name,
            user_nickname=nickname,
            details=f"인보이스 ID {invoice_id} 항목 수정, 총액: {total_amount:,}원"
        )
            
        return {
            "status": "success",
            "invoice_id": invoice_id,
            "item_count": len(request.items),
            "total_amount": total_amount,
            "modified_by": nickname
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/batch/delete")
async def delete_invoices_batch(invoice_ids: List[int], token: Optional[str] = None):
    """인보이스 일괄 삭제 (관리자만)"""
    # 관리자 권한 체크
    is_admin, nickname = check_admin(token)
    if not is_admin:
        raise HTTPException(status_code=403, detail="관리자 권한이 필요합니다.")
    
    try:
        with get_connection() as con:
            for iid in invoice_ids:
                con.execute("DELETE FROM invoice_items WHERE invoice_id = ?", (iid,))
                con.execute("DELETE FROM invoices WHERE invoice_id = ?", (iid,))
            con.commit()
        
        # 로그 기록
        add_log(
            action_type="인보이스 일괄 삭제",
            target_type="invoice",
            target_id=",".join(str(i) for i in invoice_ids),
            target_name=f"{len(invoice_ids)}건",
            user_nickname=nickname,
            details=f"인보이스 {len(invoice_ids)}건 일괄 삭제: {invoice_ids}"
        )
        
        return {"status": "success", "deleted_count": len(invoice_ids)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def _create_invoice_sheet(ws, invoice_data: dict, items_df, company_info: dict):
    """인보이스 시트 생성 헬퍼 함수 (PDF와 동일한 양식)"""
    from datetime import datetime
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    # 스타일 정의
    title_font = Font(name='맑은 고딕', size=18, bold=True)
    header_font = Font(name='맑은 고딕', size=10, bold=True)
    body_font = Font(name='맑은 고딕', size=9)
    small_font = Font(name='맑은 고딕', size=8)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    gray_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    light_gray_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # 데이터 추출
    invoice_id = invoice_data['invoice_id']
    vendor_name = invoice_data['vendor_name']
    period_from = invoice_data['period_from']
    confirmed_by = invoice_data.get('confirmed_by', '')
    
    company_name = company_info.get('company_name', '')
    business_number = company_info.get('business_number', '')
    address = company_info.get('address', '')
    business_type = company_info.get('business_type', '')
    business_item = company_info.get('business_item', '')
    bank_name = company_info.get('bank_name', '')
    account_holder = company_info.get('account_holder', '')
    account_number = company_info.get('account_number', '')
    representative = company_info.get('representative', '')
    
    # 청구일자
    invoice_date = datetime.now().strftime("%Y-%m-%d")
    
    # 건명 생성
    period_str = period_from[:7].replace("-", "년 ") + "월" if period_from else ""
    title = f"{period_str} 풀필먼트 서비스 대금"
    
    # 수신자
    recipient_name = f"{vendor_name} 대표님 귀하"
    
    # 문서번호
    doc_number = f"{invoice_id:05d}-{invoice_date.replace('-', '')[:6]}"
    
    # 지급기한
    if period_from:
        try:
            from dateutil.relativedelta import relativedelta
            period_dt = datetime.strptime(period_from[:10], "%Y-%m-%d")
            next_month = period_dt + relativedelta(months=1)
            payment_deadline = f"{next_month.year}년 {next_month.month:02d}월 05일"
        except:
            payment_deadline = ""
    else:
        payment_deadline = ""
    
    current_row = 1
    
    # 1. 제목
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'] = "물류대행 서비스 대금청구서"
    ws[f'A{current_row}'].font = title_font
    ws[f'A{current_row}'].alignment = center_align
    current_row += 2
    
    # 2. 문서번호/청구일자
    ws[f'A{current_row}'] = "문서번호"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = gray_fill
    ws[f'A{current_row}'].border = thin_border
    ws.merge_cells(f'B{current_row}:C{current_row}')
    ws[f'B{current_row}'] = doc_number
    ws[f'B{current_row}'].font = body_font
    ws[f'B{current_row}'].border = thin_border
    ws[f'D{current_row}'] = "청구일자"
    ws[f'D{current_row}'].font = header_font
    ws[f'D{current_row}'].fill = gray_fill
    ws[f'D{current_row}'].border = thin_border
    ws.merge_cells(f'E{current_row}:F{current_row}')
    ws[f'E{current_row}'] = invoice_date
    ws[f'E{current_row}'].font = body_font
    ws[f'E{current_row}'].border = thin_border
    current_row += 1
    
    # 3. 수신/건명
    ws[f'A{current_row}'] = "수신"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = gray_fill
    ws[f'A{current_row}'].border = thin_border
    ws.merge_cells(f'B{current_row}:F{current_row}')
    ws[f'B{current_row}'] = recipient_name
    ws[f'B{current_row}'].font = body_font
    ws[f'B{current_row}'].border = thin_border
    current_row += 1
    
    ws[f'A{current_row}'] = "건명"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = gray_fill
    ws[f'A{current_row}'].border = thin_border
    ws.merge_cells(f'B{current_row}:F{current_row}')
    ws[f'B{current_row}'] = title
    ws[f'B{current_row}'].font = body_font
    ws[f'B{current_row}'].border = thin_border
    current_row += 1
    
    # 4. 공급자 정보
    ws[f'A{current_row}'] = "공급자"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = gray_fill
    ws[f'A{current_row}'].border = thin_border
    ws.merge_cells(f'A{current_row}:A{current_row+2}')
    
    ws[f'B{current_row}'] = "사업자번호"
    ws[f'B{current_row}'].font = small_font
    ws[f'B{current_row}'].fill = light_gray_fill
    ws[f'B{current_row}'].border = thin_border
    ws[f'C{current_row}'] = business_number
    ws[f'C{current_row}'].font = small_font
    ws[f'C{current_row}'].border = thin_border
    ws[f'D{current_row}'] = "상호"
    ws[f'D{current_row}'].font = small_font
    ws[f'D{current_row}'].fill = light_gray_fill
    ws[f'D{current_row}'].border = thin_border
    ws.merge_cells(f'E{current_row}:F{current_row}')
    ws[f'E{current_row}'] = company_name
    ws[f'E{current_row}'].font = small_font
    ws[f'E{current_row}'].border = thin_border
    current_row += 1
    
    ws[f'B{current_row}'] = "소재지"
    ws[f'B{current_row}'].font = small_font
    ws[f'B{current_row}'].fill = light_gray_fill
    ws[f'B{current_row}'].border = thin_border
    ws.merge_cells(f'C{current_row}:F{current_row}')
    ws[f'C{current_row}'] = address
    ws[f'C{current_row}'].font = small_font
    ws[f'C{current_row}'].border = thin_border
    current_row += 1
    
    ws[f'B{current_row}'] = "업태"
    ws[f'B{current_row}'].font = small_font
    ws[f'B{current_row}'].fill = light_gray_fill
    ws[f'B{current_row}'].border = thin_border
    ws[f'C{current_row}'] = business_type
    ws[f'C{current_row}'].font = small_font
    ws[f'C{current_row}'].border = thin_border
    ws[f'D{current_row}'] = "종목"
    ws[f'D{current_row}'].font = small_font
    ws[f'D{current_row}'].fill = light_gray_fill
    ws[f'D{current_row}'].border = thin_border
    ws.merge_cells(f'E{current_row}:F{current_row}')
    ws[f'E{current_row}'] = business_item
    ws[f'E{current_row}'].font = small_font
    ws[f'E{current_row}'].border = thin_border
    current_row += 2
    
    # 5. 항목 테이블 헤더
    headers = ["No", "품명", "수량", "단가", "금액", "비고"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = gray_fill
        cell.border = thin_border
        cell.alignment = center_align
    current_row += 1
    
    # 항목 데이터
    subtotal = 0
    for idx, (_, row) in enumerate(items_df.iterrows(), 1):
        qty = int(row.get('수량', row.get('qty', 0))) if pd.notna(row.get('수량', row.get('qty'))) else 0
        unit_price = int(row.get('단가', row.get('unit_price', 0))) if pd.notna(row.get('단가', row.get('unit_price'))) else 0
        amount = int(row.get('금액', row.get('amount', qty * unit_price))) if pd.notna(row.get('금액', row.get('amount'))) else qty * unit_price
        item_name = str(row.get('항목', row.get('item_name', '')))
        remark = str(row.get('비고', row.get('remark', ''))) if pd.notna(row.get('비고', row.get('remark'))) else ""
        subtotal += amount
        
        ws.cell(row=current_row, column=1, value=idx).border = thin_border
        ws.cell(row=current_row, column=1).alignment = center_align
        ws.cell(row=current_row, column=2, value=item_name).border = thin_border
        ws.cell(row=current_row, column=3, value=f"{qty:,}" if qty else "").border = thin_border
        ws.cell(row=current_row, column=3).alignment = right_align
        ws.cell(row=current_row, column=4, value=f"{unit_price:,}" if unit_price else "").border = thin_border
        ws.cell(row=current_row, column=4).alignment = right_align
        ws.cell(row=current_row, column=5, value=f"{amount:,}" if amount else "").border = thin_border
        ws.cell(row=current_row, column=5).alignment = right_align
        ws.cell(row=current_row, column=6, value=remark).border = thin_border
        
        for col in range(1, 7):
            ws.cell(row=current_row, column=col).font = body_font
        current_row += 1
    
    current_row += 1
    
    # 6. 합계
    vat = int(subtotal * 0.1)
    total = subtotal + vat
    
    ws[f'A{current_row}'] = "합계 금액"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = gray_fill
    ws[f'A{current_row}'].border = thin_border
    ws[f'B{current_row}'] = f"₩ {subtotal:,}"
    ws[f'B{current_row}'].font = body_font
    ws[f'B{current_row}'].border = thin_border
    ws[f'B{current_row}'].alignment = right_align
    ws[f'C{current_row}'] = "부가세"
    ws[f'C{current_row}'].font = header_font
    ws[f'C{current_row}'].fill = gray_fill
    ws[f'C{current_row}'].border = thin_border
    ws[f'D{current_row}'] = f"₩ {vat:,}"
    ws[f'D{current_row}'].font = body_font
    ws[f'D{current_row}'].border = thin_border
    ws[f'D{current_row}'].alignment = right_align
    ws[f'E{current_row}'] = "청구금액"
    ws[f'E{current_row}'].font = header_font
    ws[f'E{current_row}'].fill = gray_fill
    ws[f'E{current_row}'].border = thin_border
    ws[f'F{current_row}'] = f"₩ {total:,}"
    ws[f'F{current_row}'].font = Font(name='맑은 고딕', size=11, bold=True)
    ws[f'F{current_row}'].border = thin_border
    ws[f'F{current_row}'].alignment = right_align
    current_row += 2
    
    # 7. 지급기한/계좌정보
    ws[f'A{current_row}'] = "지급기한"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = gray_fill
    ws[f'A{current_row}'].border = thin_border
    ws.merge_cells(f'B{current_row}:F{current_row}')
    ws[f'B{current_row}'] = payment_deadline
    ws[f'B{current_row}'].font = body_font
    ws[f'B{current_row}'].border = thin_border
    current_row += 1
    
    ws[f'A{current_row}'] = "계좌정보"
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].fill = gray_fill
    ws[f'A{current_row}'].border = thin_border
    ws.merge_cells(f'B{current_row}:F{current_row}')
    ws[f'B{current_row}'] = f"{bank_name}  {account_number}  {account_holder}"
    ws[f'B{current_row}'].font = body_font
    ws[f'B{current_row}'].border = thin_border
    current_row += 3
    
    # 8. 하단 - 위와 같이 청구합니다
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'] = "위와 같이 청구합니다."
    ws[f'A{current_row}'].font = header_font
    ws[f'A{current_row}'].alignment = center_align
    current_row += 2
    
    # 날짜 (한국어 형식)
    try:
        dt = datetime.strptime(invoice_date, "%Y-%m-%d")
        weekdays = ['월', '화', '수', '목', '금', '토', '일']
        date_str = f"{dt.year}년 {dt.month:02d}월 {dt.day:02d}일 {weekdays[dt.weekday()]}요일"
    except:
        date_str = invoice_date
    
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'] = date_str
    ws[f'A{current_row}'].font = body_font
    ws[f'A{current_row}'].alignment = center_align
    current_row += 2
    
    # 회사명
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'] = company_name
    ws[f'A{current_row}'].font = title_font
    ws[f'A{current_row}'].alignment = center_align
    current_row += 1
    
    # 담당자/대표자 정보
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'] = f"담당: {confirmed_by or '-'}  /  대표: {representative or '-'}"
    ws[f'A{current_row}'].font = small_font
    ws[f'A{current_row}'].alignment = center_align
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 25


@router.get("/export/xlsx")
async def export_invoices_xlsx(
    period: Optional[str] = None,
    vendor: Optional[str] = None,
    invoice_ids: Optional[str] = None  # comma-separated
):
    """인보이스 엑셀 다운로드 (PDF와 동일한 양식)"""
    try:
        from datetime import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        import re
        
        with get_connection() as con:
            # ID 목록 파싱
            ids_list = None
            if invoice_ids:
                ids_list = [int(x.strip()) for x in invoice_ids.split(',') if x.strip()]
            
            # 인보이스 컬럼 확인
            cols = [c[1] for c in con.execute("PRAGMA table_info(invoices);")]
            has_confirmed_by = 'confirmed_by' in cols
            
            # 인보이스 조회
            select_cols = """
                i.invoice_id,
                COALESCE(v.name, v.vendor, i.vendor_id) as vendor_name,
                i.period_from,
                i.period_to,
                i.total_amount,
                COALESCE(i.status, '미확정') as status
            """
            if has_confirmed_by:
                select_cols += ", i.confirmed_by"
            
            query = f"""
                SELECT {select_cols}
                FROM invoices i
                LEFT JOIN vendors v ON i.vendor_id = v.vendor_id
                WHERE 1=1
            """
            params = []
            
            if ids_list:
                placeholders = ','.join(['?' for _ in ids_list])
                query += f" AND i.invoice_id IN ({placeholders})"
                params.extend(ids_list)
            elif period:
                query += " AND strftime('%Y-%m', i.period_from) = ?"
                params.append(period)
            
            if vendor:
                query += " AND (v.vendor = ? OR v.name = ?)"
                params.extend([vendor, vendor])
            
            query += " ORDER BY i.invoice_id DESC"
            
            inv_df = pd.read_sql(query, con, params=params)
            
            if inv_df.empty:
                raise HTTPException(status_code=404, detail="No invoices found")
            
            invoice_ids_list = inv_df['invoice_id'].tolist()
            
            # 모든 항목 조회
            if invoice_ids_list:
                placeholders = ','.join(['?' for _ in invoice_ids_list])
                items_df = pd.read_sql(
                    f"SELECT invoice_id, item_name as 항목, qty as 수량, unit_price as 단가, amount as 금액, remark as 비고 FROM invoice_items WHERE invoice_id IN ({placeholders})",
                    con, params=invoice_ids_list
                )
            else:
                items_df = pd.DataFrame()
            
            # 회사 설정 조회
            company_row = con.execute("""
                SELECT company_name, business_number, address, business_type, business_item,
                       bank_name, account_holder, account_number, representative
                FROM company_settings WHERE id = 1
            """).fetchone()
        
        # 회사 정보 설정
        if company_row:
            company_info = {
                'company_name': company_row[0] or "",
                'business_number': company_row[1] or "",
                'address': company_row[2] or "",
                'business_type': company_row[3] or "",
                'business_item': company_row[4] or "",
                'bank_name': company_row[5] or "",
                'account_holder': company_row[6] or "",
                'account_number': company_row[7] or "",
                'representative': company_row[8] or "",
            }
        else:
            company_info = {
                'company_name': '', 'business_number': '', 'address': '',
                'business_type': '', 'business_item': '', 'bank_name': '',
                'account_holder': '', 'account_number': '', 'representative': ''
            }
        
        # 엑셀 워크북 생성
        wb = Workbook()
        
        # 첫 번째 시트: 인보이스 목록
        ws_list = wb.active
        ws_list.title = "목록"
        
        # 목록 스타일
        header_font = Font(name='맑은 고딕', size=10, bold=True)
        body_font = Font(name='맑은 고딕', size=9)
        gray_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # 목록 헤더
        list_headers = ["인보이스ID", "업체명", "시작일", "종료일", "총액", "상태"]
        for col_idx, header in enumerate(list_headers, 1):
            cell = ws_list.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = gray_fill
            cell.border = thin_border
        
        # 목록 데이터
        for row_idx, (_, row) in enumerate(inv_df.iterrows(), 2):
            ws_list.cell(row=row_idx, column=1, value=int(row['invoice_id'])).border = thin_border
            ws_list.cell(row=row_idx, column=2, value=str(row['vendor_name'])).border = thin_border
            ws_list.cell(row=row_idx, column=3, value=str(row['period_from']) if row['period_from'] else '').border = thin_border
            ws_list.cell(row=row_idx, column=4, value=str(row['period_to']) if row['period_to'] else '').border = thin_border
            ws_list.cell(row=row_idx, column=5, value=f"₩ {int(row['total_amount']):,}" if pd.notna(row['total_amount']) else '').border = thin_border
            ws_list.cell(row=row_idx, column=6, value=str(row['status'])).border = thin_border
            for col in range(1, 7):
                ws_list.cell(row=row_idx, column=col).font = body_font
        
        # 목록 열 너비 조정
        ws_list.column_dimensions['A'].width = 12
        ws_list.column_dimensions['B'].width = 20
        ws_list.column_dimensions['C'].width = 12
        ws_list.column_dimensions['D'].width = 12
        ws_list.column_dimensions['E'].width = 15
        ws_list.column_dimensions['F'].width = 10
        
        # 각 인보이스별 시트 (PDF 양식)
        for _, inv_row in inv_df.iterrows():
            iid = int(inv_row['invoice_id'])
            vendor_nm = str(inv_row['vendor_name'])[:15] if inv_row['vendor_name'] else 'Unknown'
            # 시트명에서 특수문자 제거
            safe_vendor = re.sub(r'[\\/*?:\[\]]', '', vendor_nm)
            sheet_name = f"{safe_vendor}_{iid}"[:31]
            
            # 새 시트 생성
            ws = wb.create_sheet(title=sheet_name)
            
            # 해당 인보이스 항목 필터링
            inv_items = items_df[items_df['invoice_id'] == iid].copy()
            
            # 인보이스 데이터 준비
            invoice_data = {
                'invoice_id': iid,
                'vendor_name': str(inv_row['vendor_name']) if inv_row['vendor_name'] else 'Unknown',
                'period_from': str(inv_row['period_from']) if inv_row['period_from'] else '',
                'confirmed_by': str(inv_row['confirmed_by']) if has_confirmed_by and pd.notna(inv_row.get('confirmed_by')) else ''
            }
            
            # 시트 생성
            _create_invoice_sheet(ws, invoice_data, inv_items, company_info)
        
        # 엑셀 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"invoices_{period if period else 'all'}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

